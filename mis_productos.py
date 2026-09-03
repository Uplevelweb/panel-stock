"""
MIS PRODUCTOS — el catálogo de ID publicados de cada cliente
=============================================================

Serling lo pidió el 01-09-2026: «debe haber dos formas de comparar, contra
nuestros ID y según lo vendido; nuestros ID preferiblemente se carguen en un
panel dentro de la sesión de cada cliente».

LAS DOS FORMAS NO SON LA MISMA PREGUNTA, Y POR ESO CONVIVEN
------------------------------------------------------------
SEGÚN LO VENDIDO   los rubros salen solos del RUT: se miran los convenios por
                   los que ya vendió y se compara contra ese mercado. Cero
                   configuración, sirve desde el primer minuto. Es lo que ve
                   alguien que acaba de escanear el QR en la feria.

CONTRA MIS ID      se cruza línea por línea contra el catálogo: de lo que esa
                   unidad compró, cuánto es de productos que él TIENE
                   PUBLICADOS. Es mucho más fino —el convenio es un saco
                   grande, el ID es el producto exacto—. El catálogo se lee
                   SOLO, del Drive: no hay nada que subir.

Lo primero contesta «¿a quién le podría vender?». Lo segundo, «¿qué de lo que
compran lo tengo yo publicado hoy?». Un vendedor necesita las dos.

DE DONDE SALE
-------------
Del Drive de ella, del archivo «CATALOGO CONVENIO MARCO», y de ningún otro
lado. Serling lo pidió así el 02-09-2026: «elimina la opción de subir el
catálogo». El catálogo no es una decisión que se tome al entrar, es un hecho.

`leer`/`guardar` siguen existiendo como red: si el Drive no contesta, se usa lo
último que hubiera quedado en la cuenta (`cuentas.ids_publicados`, que crea
`mis-productos-para-copiar.txt`) y se avisa en pantalla. Solo entonces se
dibuja algo.
"""
from __future__ import annotations

import io
import re

import pandas as pd
import streamlit as st

import modulo_cuentas

# UN ID DE CONVENIO MARCO TIENE EXACTAMENTE 7 DIGITOS. No es una suposición:
# medido el 02-09-2026 sobre la bodega entera —1.216.263 líneas, 125.874 ID
# distintos— y **el 100% tiene 7**. No hay ninguno de 6 ni de 8.
#
# Empezó siendo «5 o más» y eso metía basura: en el catálogo de verdad colaba
# 1.835 números de 5 dígitos y 272 de 6 —cantidades, códigos internos— que
# ninguno calzaba con la bodega. No producían falsos «sí lo tengo», pero sí
# inflaban la cuenta en pantalla: decía 24.763 productos cuando eran 22.656.
#
# Si algún día ChileCompra emite ID más largos, esta es la línea que se toca.
LARGO_ID = 7
CLAVE_SESION = "mis_ids"

# EL CATALOGO VIVE EN EL DRIVE DE ELLA Y SIEMPRE VA A VIVIR AHI. Lo dijo
# Serling el 02-09-2026: «ubícalo en el Drive, un archivo llamado CATALOGO
# CONVENIO MARCO, siempre reposará allí».
#
# Se baja sin credenciales por la vía de exportación de Google, la misma que ya
# usa `app.py` para la hoja de compras. El archivo está compartido por enlace;
# si algún día se deja de compartir, esto avisa y queda la carga a mano.
#
# La dirección va en el código y no en los secretos porque el repositorio ya
# lleva la carpeta de ofertas de la misma forma (`URL_OFERTAS_POR_DEFECTO`), y
# porque es el catálogo publicado en Mercado Público: es información que ella
# ya hace pública al publicarla. Si algún día quiere que no se vea, se muda a
# los secretos de Streamlit y se lee de ahí.
CATALOGO_DRIVE = "1_Z5tXDII93ovkNk-eBNFu6XIiRcbMOtb"
CATALOGO_NOMBRE = "CATALOGO CONVENIO MARCO"


# --------------------------------------------------------------------------
#  Leer el archivo que sube el cliente
# --------------------------------------------------------------------------
def ids_del_archivo(archivo) -> tuple[set[str], str]:
    """Saca los ID de un .xlsx o .csv. Devuelve (ids, explicación).

    NO SE LE PIDE AL CLIENTE QUE DIGA CUÁL ES LA COLUMNA DE ID, y es a
    propósito. El catálogo de Convenio Marco viene con una pestaña por rubro,
    con títulos y filas en blanco arriba, y la columna del ID se llama distinto
    en cada versión («ID», «ID REGIÓN CM», «ID CONVENIO REGIÓN», «ID producto»).
    Preguntar por la columna es garantizar que alguien la elija mal.

    Se leen TODAS las celdas y se toma lo que parece un ID: solo dígitos y
    exactamente 7. Con eso da igual dónde esté la columna y cómo se llame.
    """
    nombre = getattr(archivo, "name", "") or "archivo"
    try:
        contenido = archivo.read()
    except Exception:
        return set(), "No se pudo leer el archivo."

    encontrados: set[str] = set()
    cuantas_hojas = 1

    # SE LEE EN FLUJO, CELDA POR CELDA, Y NO SE ARMA NINGUNA TABLA.
    #
    # Antes esto era `pd.read_excel(sheet_name=None, dtype=str)`, que carga las
    # cuatro pestañas enteras en DataFrames: +24 MB por lectura **que no se
    # devuelven ni forzando el recolector**. Con la bodega de producción ya
    # ocupando lo suyo, unas cuantas lecturas se comían el techo de 1.000 MB de
    # Streamlit y la app moría sin traceback. Es el «Oh no» del 02-09-2026: el
    # registro mostraba decenas de lecturas seguidas del catálogo.
    #
    # Con `read_only=True` openpyxl no construye el libro en memoria: entrega
    # las filas de a una y las suelta. Y como lo único que hace falta son los
    # números de 7 dígitos, no hay por qué guardar nada más.
    if nombre.lower().endswith(".csv"):
        import csv
        for codificacion in ("utf-8-sig", "latin-1"):
            try:
                texto = contenido.decode(codificacion)
                break
            except UnicodeDecodeError:
                continue
        else:
            return set(), "No se pudo leer el archivo: codificación desconocida."
        separador = ";" if texto.count(";") >= texto.count(",") else ","
        for fila in csv.reader(io.StringIO(texto), delimiter=separador):
            for celda in fila:
                _si_es_id(celda, encontrados)
    else:
        import openpyxl
        try:
            libro = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True,
                                           data_only=True)
        except Exception as error:
            return set(), f"No se pudo abrir el archivo: {error}"
        try:
            cuantas_hojas = len(libro.sheetnames)
            for hoja in libro.worksheets:
                for fila in hoja.iter_rows(values_only=True):
                    for celda in fila:
                        _si_es_id(celda, encontrados)
        finally:
            # `read_only` deja el zip abierto: hay que cerrarlo a mano o el
            # archivo temporal queda tomado.
            libro.close()

    if not encontrados:
        return set(), ("No encontré ningún ID en ese archivo. Tienen que ser "
                       f"números de {LARGO_ID} dígitos, en cualquier columna.")
    return encontrados, (f"{len(encontrados):,}".replace(",", ".") +
                         f" productos leídos de «{nombre}»" +
                         (f", {cuantas_hojas} pestañas" if cuantas_hojas > 1 else ""))


def _si_es_id(celda, encontrados: set) -> None:
    """Guarda la celda si parece un ID de Convenio Marco.

    ⚠️ LA MITAD DE LOS ID VIENEN COMO NUMERO, NO COMO TEXTO. openpyxl los
    entrega tal cual están guardados, así que un ID numérico llega como
    `4194137.0` y `str()` lo deja de 9 caracteres: se descartaba. Con
    `pd.read_excel(dtype=str)` esto no se notaba porque pandas los convertía.
    Costó 10.346 productos de 22.656 en la primera versión de este lector.

    Acepta también el número escrito con puntos de miles («4.194.137»), que es
    como Excel a veces deja la columna.
    """
    if celda is None:
        return
    if isinstance(celda, bool):
        return
    if isinstance(celda, int):
        texto = str(celda)
    elif isinstance(celda, float):
        # Un ID no tiene decimales: si los trae, no es un ID.
        if not celda.is_integer():
            return
        texto = str(int(celda))
    else:
        texto = str(celda).strip().replace(".", "")
    if len(texto) == LARGO_ID and texto.isdigit():
        encontrados.add(texto)


# --------------------------------------------------------------------------
#  El catálogo del Drive
# --------------------------------------------------------------------------
@st.cache_data(ttl=600, show_spinner="Leyendo tu catálogo del Drive…")
def ids_del_drive() -> tuple[set, str]:
    """Baja «CATALOGO CONVENIO MARCO» del Drive y saca sus ID.

    Diez minutos de cache: es el mismo plazo que usa `app.cargar_catalogo_propio`
    para el otro catálogo. Suficiente para que ella actualice el archivo y lo
    vea reflejado en la siguiente vuelta, sin bajarlo en cada clic.
    """
    import urllib.error
    import urllib.request

    url = (f"https://docs.google.com/spreadsheets/d/{CATALOGO_DRIVE}"
           "/export?format=xlsx")
    try:
        peticion = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(peticion, timeout=90) as respuesta:
            datos = respuesta.read()
    except urllib.error.HTTPError as error:
        return set(), (f"El Drive respondió {error.code}. Puede que el archivo "
                       "haya dejado de estar compartido por enlace.")
    except Exception as error:
        return set(), f"No se pudo bajar del Drive: {type(error).__name__}."

    # Si Google devuelve una página en vez del archivo, no es un xlsx: pasa
    # cuando el enlace dejó de ser público y contesta con el formulario de
    # acceso. Sin esta comprobación, el lector de más abajo diría «no encontré
    # ningún ID», que manda a buscar el problema al lado equivocado.
    if datos[:4] != b"PK\x03\x04":
        return set(), ("El Drive devolvió una página en vez del archivo: "
                       "seguramente dejó de estar compartido por enlace.")

    envoltorio = io.BytesIO(datos)
    envoltorio.name = f"{CATALOGO_NOMBRE}.xlsx"
    return ids_del_archivo(envoltorio)


# --------------------------------------------------------------------------
#  Guardar y recuperar
# --------------------------------------------------------------------------
def _cuenta_de(usuario: dict) -> str:
    return str((usuario or {}).get("cuenta_id") or "")


def leer(usuario: dict) -> set[str]:
    """Los ID de esa empresa. Primero la cuenta; si no, lo de la sesión."""
    cuenta = _cuenta_de(usuario)
    if cuenta:
        filas = modulo_cuentas._pedir(
            f"cuentas?select=ids_publicados&id=eq.{cuenta}&limit=1")
        if filas:
            guardados = (filas[0] or {}).get("ids_publicados")
            if guardados:
                return {str(x) for x in guardados}
    return set(st.session_state.get(CLAVE_SESION) or [])


def guardar(usuario: dict, ids: set[str]) -> tuple[bool, str]:
    """Deja los ID en la cuenta. Si la columna no existe, en la sesión."""
    # Siempre en la sesión: es lo que hace que funcione en el momento, aunque
    # la base conteste mal.
    st.session_state[CLAVE_SESION] = sorted(ids)

    cuenta = _cuenta_de(usuario)
    if not cuenta:
        return True, "Cargados para esta sesión."

    respuesta = modulo_cuentas._pedir(
        f"cuentas?id=eq.{cuenta}", "PATCH",
        {"ids_publicados": sorted(ids)},
        extra={"Prefer": "return=representation"})
    if respuesta is None:
        return True, ("Cargados **para esta sesión**. Para que queden guardados "
                      "hay que correr `mis-productos-para-copiar.txt` en Supabase.")
    return True, "Guardados en tu cuenta."


def borrar(usuario: dict) -> None:
    st.session_state.pop(CLAVE_SESION, None)
    cuenta = _cuenta_de(usuario)
    if cuenta:
        modulo_cuentas._pedir(f"cuentas?id=eq.{cuenta}", "PATCH",
                              {"ids_publicados": None})


# --------------------------------------------------------------------------
#  La pantalla
# --------------------------------------------------------------------------
def ids_vigentes(usuario: dict) -> set[str]:
    """El catálogo del cliente. Devuelve los ID y NO dibuja nada si todo va bien.

    SE LEE SOLO DEL DRIVE, Y EN SILENCIO. Serling lo pidió así el 02-09-2026:
    «elimina la opción de subir el catálogo, no debería aparecer el link de mis
    productos publicados». Y tiene razón: el catálogo no es una decisión que
    ella tome cada vez que entra, es un hecho. Un panel que se abre pidiendo
    algo que ya está resuelto le hace perder tiempo a todo el mundo.

    Solo se dibuja algo cuando FALLA, que es cuando hace falta enterarse.
    """
    del_drive, aviso_drive = ids_del_drive()
    if del_drive:
        return del_drive

    # Solo si el Drive fallo se dice algo, y se dice donde se esta mirando.
    guardados = leer(usuario)
    st.warning(
        f"No se pudo leer **{CATALOGO_NOMBRE}** del Drive. {aviso_drive}"
        + (f" Se usan los {len(guardados):,}".replace(",", ".") +
           " productos que quedaron de antes."
           if guardados else
           " Mientras tanto, «Contra mis ID publicados» no puede calcularse: "
           "usa «Segun lo que ya has vendido», que no necesita catalogo."))
    return guardados
